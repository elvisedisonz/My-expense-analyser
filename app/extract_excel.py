import io
import warnings

import msoffcrypto
import pandas as pd

from openpyxl import load_workbook


HEADER_COLUMNS = [
    "Date",
    "Details",
    "Debit",
    "Credit",
    "Balance",
]

END_MARKER = "Statement Summary"

STANDARD_COLUMNS = [
    "Date",
    "Details",
    "Debit",
    "Credit",
    "Balance",
]


def normalize_dataframe(df):

    if df is None or df.empty:
        return None

    df.columns = [
        str(c).strip()
        for c in df.columns
    ]

    keep = [
        c for c in STANDARD_COLUMNS
        if c in df.columns
    ]

    df = df[keep]

    df = df.dropna(how="all")

    return df.reset_index(drop=True)


def row_matches_header(values):

    normalized = [
        str(v).strip()
        for v in values
    ]

    return all(
        col in normalized
        for col in HEADER_COLUMNS
    )


def load_excel_workbook(
    file_path,
    password=None
):

    # ---------------- TRY DECRYPT ----------------

    try:

        with open(file_path, "rb") as f:

            office_file = msoffcrypto.OfficeFile(f)

            if password:
                office_file.load_key(password=password)
            else:
                office_file.load_key(password="")

            decrypted = io.BytesIO()

            office_file.decrypt(decrypted)

        decrypted.seek(0)

        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style"
        )

        return load_workbook(decrypted)

    # ---------------- NORMAL EXCEL ----------------

    except Exception:

        return load_workbook(file_path)


def excel_to_dataframe(
    file_path,
    password=None
):

    try:

        workbook = load_excel_workbook(
            file_path,
            password=password
        )

        target_sheet = None

        for sheet in workbook.worksheets:

            start_row = None
            end_row = None

            # ---------------- UNMERGE ----------------

            for merged in list(
                sheet.merged_cells.ranges
            ):
                sheet.unmerge_cells(str(merged))

            # ---------------- FIND HEADER ----------------

            for row in sheet.iter_rows():

                values = [
                    str(cell.value).strip()
                    if cell.value is not None else ""
                    for cell in row
                ]

                if row_matches_header(values):
                    start_row = row[0].row
                    break

            if start_row is None:
                continue

            # ---------------- FIND END ----------------

            for row in sheet.iter_rows(
                min_row=start_row + 1
            ):

                row_values = [
                    str(cell.value)
                    for cell in row
                    if cell.value is not None
                ]

                if any(
                    END_MARKER.lower()
                    in value.lower()
                    for value in row_values
                ):
                    end_row = row[0].row
                    break

            # ---------------- REMOVE TOP ----------------

            if start_row > 1:

                sheet.delete_rows(
                    1,
                    start_row - 1
                )

            # ---------------- REMOVE BOTTOM ----------------

            if end_row is not None:

                adjusted_end = (
                    end_row - start_row + 1
                )

                sheet.delete_rows(
                    adjusted_end,
                    sheet.max_row - adjusted_end + 1
                )

            target_sheet = sheet
            break

        if target_sheet is None:
            return None

        # ---------------- DATAFRAME ----------------

        data = target_sheet.values

        columns = next(data)

        df = pd.DataFrame(
            data,
            columns=columns
        )

        df = normalize_dataframe(df)

        # ---------------- SAVE ----------------
        if df is not None and not df.empty:
            output_path = "debug_excel.xlsx"
            df.to_excel(output_path, index=False)

        return df

    except Exception as e:

        raise Exception(
            f"Excel extraction failed: {e}"
        )
